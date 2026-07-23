"""
fcs_export.py
=============
Shared CSV export for the FCS analysis suite.

When the "Export plotted data to CSV" option is ticked in the main window,
each analysis writes the exact arrays it plots to a CSV file inside an
``analysis`` folder created alongside the source ``.fcs`` file.

File layout
-----------
Each export is a plain CSV with a commented metadata block on top::

    # FCS analysis export — intensity
    # source file : experiment.fcs
    # exported    : 2026-06-10T17:30:00
    # bin_width_s : 0.01
    time_s,ch1_counts,ch2_counts,ch1_cps,ch2_cps
    0,12,9,1200,900
    ...

The ``#`` lines are comments (skip them with ``pandas.read_csv(path,
comment='#')`` or ``numpy.loadtxt`` defaults).  The first non-comment line
is the real column header, followed by the data.

Filenames
---------
``<source-stem>_<analysis>[_<suffix>].csv``

The optional suffix captures the primary plot variant (e.g. lifetime bin
count, correlation type, PCH channel + bin width) so that two genuinely
different plots of the same file do not overwrite one another, while
re-running the *identical* plot overwrites its previous export.

Set ``TIMESTAMP_FILENAMES = True`` to instead keep every export by
appending a timestamp to the filename.

Public API
----------
    analysis_dir(d)                                  -> Path
    export_columns(d, analysis, columns, meta, ...)  -> Path   (raises)
    safe_export(d, analysis, columns, meta, ...)     -> Path | None
    read_export(path)                                -> (meta, columns)
    write_table_xlsx(path, comments, header, rows)   -> Path | None
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime
from pathlib import Path
from typing import Mapping, Optional, Tuple

import numpy as np

from fcs_reader import FCSData


# If True, a timestamp is inserted into each filename so repeated exports are
# all preserved.  If False (default), re-running an analysis with the same
# settings overwrites that analysis's previous export for the same file.
TIMESTAMP_FILENAMES = False

_OUTPUT_DIRNAME = "analysis"


# ── Folder ────────────────────────────────────────────────────────────────────

def analysis_dir(d: FCSData) -> Path:
    """
    Return the ``analysis`` folder beside the source file, creating it
    (and any missing parents) if it does not already exist.
    """
    out = d.filepath.parent / _OUTPUT_DIRNAME
    out.mkdir(parents=True, exist_ok=True)
    return out


# ── Filename helper ───────────────────────────────────────────────────────────

def _slug(text: str) -> str:
    """Reduce arbitrary text to a filesystem-safe lower-case token."""
    out = []
    for ch in str(text).strip().lower():
        out.append(ch if ch.isalnum() else "_")
    slug = "".join(out).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug or "data"


# ── Core writer ───────────────────────────────────────────────────────────────

def export_columns(
    d: FCSData,
    analysis: str,
    columns: Mapping[str, np.ndarray],
    meta: Optional[Mapping[str, object]] = None,
    *,
    suffix: str = "",
) -> Path:
    """
    Write a set of named, equal-length columns to a CSV in the analysis folder.

    Parameters
    ----------
    d        : FCSData
        Source dataset — provides the output folder, filename stem and the
        file-name fields written into the header.
    analysis : str
        Short analysis name, e.g. ``"intensity"`` or ``"correlation"``.
    columns  : mapping of column-label -> 1-D array
        All arrays must share the same length; they become the CSV columns
        in insertion order.  The label is used verbatim as the column header.
    meta     : mapping, optional
        Extra ``key : value`` lines written into the commented header block
        (parameters, statistics, photon counts, …).
    suffix   : str, keyword-only
        Extra tag folded into the filename to distinguish plot variants.

    Returns
    -------
    Path to the written file.

    Raises
    ------
    ValueError
        If no columns are supplied or their lengths differ.
    """
    names = list(columns.keys())
    if not names:
        raise ValueError("No columns supplied to export.")

    arrays = [np.asarray(columns[k], dtype=np.float64).ravel() for k in names]
    n = len(arrays[0])
    for name, arr in zip(names, arrays):
        if len(arr) != n:
            raise ValueError(
                f"Column '{name}' has length {len(arr)}, expected {n}; "
                f"all export columns must be the same length."
            )

    # ── Build the output path ────────────────────────────────────────────────
    parts = [_slug(d.filepath.stem), _slug(analysis)]
    if suffix:
        parts.append(_slug(suffix))
    if TIMESTAMP_FILENAMES:
        parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    out_path = analysis_dir(d) / ("_".join(parts) + ".csv")

    # ── Write commented header + real CSV header + data ──────────────────────
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        fh.write(f"# FCS analysis export \u2014 {analysis}\n")
        # The banner above is for a human opening the file in a text editor.
        # This key is the machine-readable one: the banner uses an em-dash and
        # no colon, so a "key : value" parser skips it entirely and cannot
        # tell what analysis produced the file.
        fh.write(f"# analysis    : {analysis}\n")
        fh.write(f"# source file : {d.filepath.name}\n")
        fh.write(f"# exported    : {datetime.now().isoformat(timespec='seconds')}\n")
        if meta:
            for key, val in meta.items():
                fh.write(f"# {key} : {val}\n")
        fh.write(",".join(names) + "\n")
        for row in zip(*arrays):
            fh.write(",".join(f"{v:.10g}" for v in row) + "\n")

    return out_path


# ── Core reader ───────────────────────────────────────────────────────────────

def read_export(path) -> Tuple[dict, dict]:
    """
    Read a CSV written by :func:`export_columns` back into (meta, columns).

    This is the inverse of the writer and the single place the header format
    is parsed.  Several modules previously each carried their own near-
    identical parser; they should migrate to this one so the format is
    defined once.

    Parsing rules
    -------------
    * ``# key : value`` lines become string entries in *meta*.  Only the FIRST
      colon separates, so values containing colons (timestamps, prose
      definitions) survive intact.
    * The legacy ``# FCS analysis export — <name>`` banner is recognised too,
      so files written before the ``analysis`` key existed still identify
      themselves.  An explicit ``analysis`` key always wins.
    * The first non-comment line is the column header; the rest is data.
    * A column parses to a float array when every value is numeric, and to an
      object array of strings otherwise (dataset names, yes/no flags).

    Parameters
    ----------
    path : str | Path
        CSV to read.

    Returns
    -------
    (meta, columns)
        *meta* maps header key -> string value.
        *columns* maps column name -> np.ndarray, in file order.

    Raises
    ------
    ValueError
        If the file contains no column header.
    """
    meta: dict = {}
    header = None
    rows: list = []

    with Path(path).open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        for raw in reader:
            if not raw:
                continue
            first = raw[0]
            if first.lstrip().startswith("#"):
                # Comment rows may have been split on commas by the reader;
                # rejoin so values containing commas are not truncated.
                body = ",".join(raw).lstrip()[1:].strip()
                if ":" in body:
                    key, val = body.split(":", 1)
                    meta[key.strip()] = val.strip()
                elif "\u2014" in body:
                    # Legacy banners, for files written before the analysis
                    # key existed.  Matched on the text LEFT of the dash --
                    # the fit banner is "FCS global fit curves — <model key>",
                    # so splitting blindly would record the model key as the
                    # analysis name.
                    left, right = body.split("\u2014", 1)
                    banner, tail = left.strip().lower(), right.strip()
                    if banner == "fcs analysis export":
                        meta.setdefault("analysis", tail)
                    elif banner == "fcs global fit curves":
                        meta.setdefault("analysis", "global fit curves")
                        meta.setdefault("model_key", tail)
                continue
            if header is None:
                header = [c.strip() for c in raw]
                continue
            rows.append(raw)

    if header is None:
        raise ValueError(f"No column header found in {path}; "
                         f"this does not look like an FCS analysis export.")

    columns: dict = {}
    for i, name in enumerate(header):
        cell = [(r[i] if i < len(r) else "") for r in rows]
        try:
            columns[name] = np.array(
                [float(c) if c.strip() != "" else np.nan for c in cell],
                dtype=float)
        except ValueError:
            columns[name] = np.array([c.strip() for c in cell], dtype=object)

    return meta, columns


# ── Spreadsheet mirror ────────────────────────────────────────────────────────

def write_table_xlsx(
    path,
    comments: list,
    header: list,
    rows: list,
    *,
    sheet_title: str = "data",
    log_tag: str = "export",
):
    """
    Mirror a table as a real .xlsx for convenient reading in Excel.

    Numbers are written as numbers, so Excel sorts and formats them natively
    and there is no text-encoding step to produce mojibake.  The CSV remains
    the machine-readable copy that the rest of the suite reads; this is purely
    for the human.

    Requires openpyxl.  If it is missing, or the write fails, this is a no-op
    that returns None -- an unwritable spreadsheet must never cost you the CSV
    that was already written.

    Parameters
    ----------
    path : str | Path
        Destination .xlsx.
    comments : list of str
        Header block, written in grey italics above the table.
    header : list of str
        Column titles, written bold and frozen.
    rows : list of sequences
        Table body.  Non-finite floats become blank cells rather than the
        string "nan", which Excel would treat as text and refuse to plot.
    sheet_title : str
        Worksheet name.
    log_tag : str
        Prefix for the console messages, e.g. "globalfit" or "calib".
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"[{log_tag}] openpyxl not installed — wrote .csv only "
              f"(pip install openpyxl to also get .xlsx).")
        return None

    import math
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    grey = Font(color="808080", italic=True)
    bold = Font(bold=True)

    r = 1
    for line in comments:
        ws.cell(row=r, column=1, value=line).font = grey
        r += 1
    for j, h in enumerate(header, start=1):
        ws.cell(row=r, column=j, value=h).font = bold
    header_row = r
    r += 1
    for row in rows:
        for j, v in enumerate(row, start=1):
            if isinstance(v, float) and not math.isfinite(v):
                v = None                              # NaN/inf -> blank cell
            ws.cell(row=r, column=j, value=v)
        r += 1

    for j, h in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(
            len(str(h)) + 2, 12)
    ws.freeze_panes = f"A{header_row + 1}"

    try:
        wb.save(path)
    except Exception as e:
        print(f"[{log_tag}] could not write .xlsx ({e}); .csv is unaffected.")
        return None
    print(f"[{log_tag}] wrote {path}")
    return Path(path)


def safe_export(
    d: FCSData,
    analysis: str,
    columns: Mapping[str, np.ndarray],
    meta: Optional[Mapping[str, object]] = None,
    *,
    suffix: str = "",
) -> Optional[Path]:
    """
    Like :func:`export_columns` but never raises.

    A failed export (e.g. a read-only directory) prints a warning and
    returns ``None`` so the plot still appears.  On success the path is
    printed and returned.
    """
    try:
        path = export_columns(d, analysis, columns, meta=meta, suffix=suffix)
        print(f"[export] wrote {path}")
        return path
    except Exception as exc:  # noqa: BLE001 — export must never break a plot
        print(f"[export] FAILED for {analysis}: {exc}", file=sys.stderr)
        return None
