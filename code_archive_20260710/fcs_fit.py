"""
fcs_fit.py
==========
Fit saved correlation curves to an FCS model.

Workflow (launched from the main window)
----------------------------------------
    1. Pick a correlation CSV  (defaults to the active file's analysis folder;
       these are the files written by the "Export plotted data to CSV" option).
    2. Choose a model               -> _select_model_dialog
    3. Set guesses / bounds / fixed -> _global_setup_dialog
    4. Fit, plot data + fit + residuals, and write results to a 'fits' folder.

The numerical core (load_correlation_csv, auto_guess, fit_correlation) has no
GUI dependency and can be reused for batch / multi-dataset fitting later.

Models come from fcs_models.MODELS — see that file to add or edit models.

Dependencies
------------
    pip install numpy scipy matplotlib
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import fcs_plottools
from scipy.optimize import curve_fit, least_squares

import fcs_models
from fcs_models import FCSModel
from fcs_reader import read_fcs, FCSData
import fcs_lifetime
import fcs_lifetime_fit
import fcs_pch_fit
import fcs_fisher
import fcs_noise

# ── CSV loading ───────────────────────────────────────────────────────────────

def load_correlation_csv(
    path: str | Path,
) -> Tuple[np.ndarray, np.ndarray, Optional[np.ndarray], Dict[str, np.ndarray], Dict[str, str]]:
    """
    Read a correlation export written by fcs_corr / fcs_export.

    Skips ``#`` comment lines (parsing ``# key : value`` lines into a metadata
    dict), reads the column header, and returns the lag axis in seconds, the
    correlation G, and (if present) the per-segment standard deviation G_std.

    Returns
    -------
    tau_s : np.ndarray
    G     : np.ndarray
    G_std : np.ndarray or None
    columns : dict of every column found, by name
    meta  : dict of header ``# key : value`` fields (e.g. 'source file', 'type')
    """
    path = Path(path)
    header: Optional[list[str]] = None
    rows: list[list[str]] = []
    meta: Dict[str, str] = {}
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            if line.startswith("#"):
                body = line[1:].strip()
                if ":" in body:
                    k, v = body.split(":", 1)
                    meta[k.strip()] = v.strip()
                continue
            if not line.strip():
                continue
            cells = [c.strip() for c in line.rstrip("\n").split(",")]
            if header is None:
                header = cells
                continue
            rows.append(cells)

    if not header:
        raise ValueError(f"No header row found in {path.name}.")

    data = (np.array(rows, dtype=np.float64)
            if rows else np.empty((0, len(header)), dtype=np.float64))
    columns = {name: data[:, i] for i, name in enumerate(header)}

    if "tau_s" in columns:
        tau_s = columns["tau_s"]
    elif "tau_ms" in columns:
        tau_s = columns["tau_ms"] * 1e-3
    else:
        raise ValueError(
            f"{path.name} has no 'tau_s' or 'tau_ms' column — "
            f"is this a correlation export?"
        )

    if "G" not in columns:
        raise ValueError(
            f"{path.name} has no 'G' column — is this a correlation export?"
        )
    G = columns["G"]
    G_std = columns.get("G_std")

    return tau_s, G, G_std, columns, meta


def _cps_from_meta(meta: Optional[dict]) -> Optional[dict]:
    """
    Derive measurement CPS (and acquisition time / photon counts) from a
    correlation CSV's header meta.

    The single "fit CPS" follows the correlation type: for an autocorrelation
    it is that channel's rate; for a cross-correlation it is the average of the
    two channels (not a fitted quantity, but a convenient brightness summary).

    Returns a dict with floats (None where a field is absent), or None when the
    header carries no CPS at all — e.g. a correlation file written before this
    field existed — so callers can simply skip the section.
    """
    if not meta:
        return None

    def _num(key):
        v = meta.get(key)
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    cps1 = _num("cps_ch1")
    cps2 = _num("cps_ch2")
    if cps1 is None and cps2 is None:
        return None   # not a CPS-aware correlation export

    ctype = (meta.get("type") or "cross").strip()
    if ctype == "auto_ch1":
        cps_fit, label = cps1, "Ch1"
    elif ctype == "auto_ch2":
        cps_fit, label = cps2, "Ch2"
    else:   # cross (or unknown) → average of the available channels
        present = [c for c in (cps1, cps2) if c is not None]
        cps_fit = (sum(present) / len(present)) if present else None
        label = "mean(Ch1, Ch2)"

    return {
        "cps_ch1":       cps1,
        "cps_ch2":       cps2,
        "cps_fit":       cps_fit,
        "cps_label":     label,
        "acq_time_s":    _num("acquisition_time_s"),
        "n_photons_ch1": _num("n_photons_ch1"),
        "n_photons_ch2": _num("n_photons_ch2"),
    }


def _fmt_cps(x: Optional[float]) -> str:
    """Format a CPS / photon count for a report (thousands-separated)."""
    if x is None or (isinstance(x, float) and not np.isfinite(x)):
        return "n/a"
    return f"{x:,.0f}"


# ── Initial-guess heuristics ──────────────────────────────────────────────────

def auto_guess(model: FCSModel, tau_s: np.ndarray, G: np.ndarray) -> Dict[str, float]:
    """
    Produce sensible starting guesses from the data.

    Generic models fall back to their declared defaults; for the standard
    diffusion parameters (G0, tau_D, offset) the amplitude is read from the
    short-lag plateau and the diffusion time from the half-amplitude crossing.
    """
    guess = model.defaults()

    t = np.asarray(tau_s, dtype=np.float64)
    y = np.asarray(G, dtype=np.float64)
    m = np.isfinite(t) & np.isfinite(y) & (t > 0)
    t, y = t[m], y[m]
    if len(t) < 3:
        return guess

    order = np.argsort(t)
    t, y = t[order], y[order]

    # Amplitude from the first few (shortest-lag) finite points.
    n_head = max(3, len(y) // 20)
    amp = float(np.median(y[:n_head]))

    if "offset" in guess:
        guess["offset"] = 0.0
    if "G0" in guess:
        guess["G0"] = max(amp, 1e-9)

    # Diffusion time: first lag where G falls below half amplitude.
    if "tau_D" in guess:
        off    = guess.get("offset", 0.0)
        amp0   = guess.get("G0", amp)
        target = off + 0.5 * amp0
        below  = np.where(y < target)[0]
        if len(below):
            guess["tau_D"] = float(t[below[0]])
        else:
            guess["tau_D"] = float(np.sqrt(t[0] * t[-1]))   # geometric midpoint

    return guess



# ── Export ────────────────────────────────────────────────────────────────────

def _fits_dir(source_path: Path) -> Path:
    """
    Return (creating if needed) a 'fits' folder beside the original data file.

    Correlation CSVs normally live in '<datadir>/analysis/', so the fits
    folder is placed as a sibling: '<datadir>/fits/'.  If the source is not
    inside an 'analysis' folder, 'fits' is created next to it instead.
    """
    base = source_path.parent
    if base.name.lower() == "analysis":
        base = base.parent
    out = base / "fits"
    out.mkdir(parents=True, exist_ok=True)
    return out

def _new_fit_dir(source_path: Path) -> Path:
    """
    Create and return a fresh, uniquely-named subfolder inside the shared
    'fits' directory for ONE export.  The folder name is just the export
    timestamp; sample, model and parameter detail lives inside the files.
    """
    fits = _fits_dir(source_path)                    # the shared 'fits' dir
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out = fits / stamp
    n = 2
    while out.exists():                              # avoid same-second clashes
        out = fits / f"{stamp}_{n}"
        n += 1
    out.mkdir(parents=True, exist_ok=True)
    return out



def _write_params_xlsx(path: Path, comments: list, header: list,
                        rows: list) -> Optional[Path]:
    """
    Mirror the parameter table as a real .xlsx for convenient viewing in Excel.
    Numbers are written as numbers (so Excel sorts/formats them natively) and no
    text-encoding step is involved, so there's no mojibake.  The .csv remains
    the machine-readable copy that 'Calibrate Volume' reads; this .xlsx is
    purely for the human.  Requires openpyxl; if it's missing this is a no-op
    and the .csv is unaffected.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[globalfit] openpyxl not installed — wrote .csv only "
              "(pip install openpyxl to also get .xlsx).")
        return None

    import math
    wb = Workbook()
    ws = wb.active
    ws.title = "fit parameters"
    grey = Font(color="808080", italic=True)
    bold = Font(bold=True)

    r = 1
    for line in comments:
        ws.cell(row=r, column=1, value=line).font = grey
        r += 1
    for j, h in enumerate(header, start=1):
        ws.cell(row=r, column=j, value=h).font = bold
    header_row = r
    r += 1
    for row in rows:
        for j, v in enumerate(row, start=1):
            if isinstance(v, float) and not math.isfinite(v):
                v = None                              # NaN/inf -> blank cell
            ws.cell(row=r, column=j, value=v)
        r += 1

    for j, h in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(len(str(h)) + 2, 12)
    ws.freeze_panes = f"A{header_row + 1}"

    try:
        wb.save(path)
    except Exception as e:
        print(f"[globalfit] could not write .xlsx ({e}); .csv is unaffected.")
        return None
    print(f"[globalfit] wrote {path}")
    return path



def _fmt(x: float) -> str:
    if x == np.inf:
        return "inf"
    if x == -np.inf:
        return "-inf"
    return f"{x:.4g}"


# ── Global / linked fitting ───────────────────────────────────────────────────

def combined_guess(model: FCSModel, datasets: list) -> Dict[str, float]:
    """
    A shared starting guess across datasets: the median of each dataset's own
    auto-guess (falls back to model defaults where data are unusable).
    """
    g = model.defaults()
    per = [auto_guess(model, ds["tau"], ds["G"]) for ds in datasets]
    for n in g:
        vals = [p[n] for p in per if np.isfinite(p.get(n, np.nan))]
        if vals:
            g[n] = float(np.median(vals))
    return g


def fit_global(
    model: FCSModel,
    datasets: list,
    linked: Dict[str, bool],
    guesses: Dict[str, float],
    lowers: Dict[str, float],
    uppers: Dict[str, float],
    fixed: Dict[str, bool],
    weighted: bool = False,
    maxfev: int = 20000,
) -> dict:
    """
    Global least-squares fit of one model to several correlation datasets.

    Each model parameter is one of:
      * fixed   — held at its guess for every dataset (no free variable);
      * linked  — a single shared free variable used by every dataset;
      * unlinked— one independent free variable per dataset.

    ``datasets`` is a list of dicts with keys ``name``, ``tau``, ``G`` and
    optionally ``sigma``.  Weighting is applied only if ``weighted`` is True
    *and* every included dataset carries a usable sigma.

    Returns a result dict with global goodness-of-fit plus a per-dataset
    breakdown (values, 1σ errors, fit curve, residuals, R²).
    """
    names = model.param_names()

    # ── Mask each dataset to finite, positive-lag points ─────────────────────
    prepped = []
    for ds in datasets:
        t = np.asarray(ds["tau"], dtype=np.float64)
        y = np.asarray(ds["G"],   dtype=np.float64)
        m = np.isfinite(t) & np.isfinite(y) & (t > 0)
        s = None
        if weighted and ds.get("sigma") is not None:
            s = np.asarray(ds["sigma"], dtype=np.float64)
            m &= np.isfinite(s) & (s > 0)
        t, y = t[m], y[m]
        s = s[m] if s is not None else None
        if len(t) < 2:
            raise ValueError(f"Dataset '{ds['name']}' has too few finite points to fit.")
        prepped.append({"name": ds["name"], "tau": t, "G": y, "sigma": s,
                        "meta": ds.get("meta", {})})

    D = len(prepped)
    if D == 0:
        raise ValueError("No datasets selected.")
    use_weights = weighted and all(p["sigma"] is not None for p in prepped)

    # ── Lay out the free-parameter vector ────────────────────────────────────
    free_spec: list = []          # (param_name, dataset_index | None)
    for p in names:
        if fixed.get(p, False):
            continue
        if linked.get(p, False):
            free_spec.append((p, None))
        else:
            free_spec.extend((p, di) for di in range(D))
    if not free_spec:
        raise ValueError("At least one parameter must be free (not fixed).")
    idx = {spec: i for i, spec in enumerate(free_spec)}
    fixed_value = {p: guesses[p] for p in names if fixed.get(p, False)}

    def value_of(theta, p, di):
        if fixed.get(p, False):
            return fixed_value[p]
        if linked.get(p, False):
            return theta[idx[(p, None)]]
        return theta[idx[(p, di)]]

    def residuals(theta):
        chunks = []
        for di, pp in enumerate(prepped):
            vals = {p: value_of(theta, p, di) for p in names}
            mvals = model.func(pp["tau"], **vals)
            r = pp["G"] - mvals
            if use_weights:
                r = r / pp["sigma"]
            chunks.append(r)
        return np.concatenate(chunks)

    theta0, lb, ub = [], [], []
    for (p, _di) in free_spec:
        theta0.append(guesses[p]); lb.append(lowers[p]); ub.append(uppers[p])
    theta0 = [min(max(v, lo), hi) for v, lo, hi in zip(theta0, lb, ub)]

    sol = least_squares(residuals, theta0, bounds=(lb, ub), max_nfev=maxfev)

    # ── Covariance / parameter errors ────────────────────────────────────────
    n_obs = int(sum(len(pp["tau"]) for pp in prepped))
    n_par = len(free_spec)
    dof = n_obs - n_par
    cov = None
    try:
        JtJ = sol.jac.T @ sol.jac
        cov = np.linalg.pinv(JtJ)
        if not use_weights and dof > 0:
            cov = cov * (2.0 * sol.cost / dof)   # scale by residual variance
        perr_free = np.sqrt(np.clip(np.diag(cov), 0.0, np.inf))
    except Exception:
        perr_free = np.full(n_par, np.nan)

    def err_of(p, di):
        if fixed.get(p, False):
            return 0.0
        if linked.get(p, False):
            return float(perr_free[idx[(p, None)]])
        return float(perr_free[idx[(p, di)]])

    # ── Per-dataset breakdown ────────────────────────────────────────────────
    theta = sol.x
    per_dataset = []
    for di, pp in enumerate(prepped):
        vals = {p: value_of(theta, p, di) for p in names}
        errs = {p: err_of(p, di) for p in names}
        mvals = model.func(pp["tau"], **vals)
        resid = pp["G"] - mvals
        ss_res = float(np.sum(resid ** 2))
        ss_tot = float(np.sum((pp["G"] - pp["G"].mean()) ** 2))
        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")
        per_dataset.append({
            "name": pp["name"], "tau": pp["tau"], "G": pp["G"],
            "Gfit": mvals, "resid": resid, "sigma": pp["sigma"],
            "values": vals, "errors": errs, "r2": r2,
            "ss_res": ss_res, "n_points": len(pp["tau"]),
            "meta": pp.get("meta", {}),
        })

    full_res = residuals(theta)
    ss_res_tot = float(np.sum(full_res ** 2))
    if use_weights and dof > 0:
        chi2, red_chi2 = ss_res_tot, ss_res_tot / dof
    else:
        chi2 = red_chi2 = float("nan")

    #this builds labels for the free param vector
    free_labels = [
        p if di is None else f"{p}[{prepped[di]['name']}]"
        for (p, di) in free_spec
    ]

    return {
        "model": model, "names": names, "datasets": per_dataset,
        "linked": dict(linked), "fixed": dict(fixed),
        "lowers": dict(lowers), "uppers": dict(uppers), "guesses": dict(guesses),
        "weighted": use_weights, "n_datasets": D,
        "dof": dof, "n_free": n_par, "n_obs": n_obs,
        "chi2": chi2, "red_chi2": red_chi2, "ss_res": ss_res_tot,
        "success": bool(sol.success), "message": str(sol.message),
        "cov": cov, "free_labels": free_labels,
    }


def plot_global_fit(result: dict, show: bool = True):
    """Overlay every dataset's data + fit, with a shared residual panel."""
    model = result["model"]
    dsets = result["datasets"]
    cmap = plt.cm.tab10 if len(dsets) <= 10 else plt.cm.viridis
    colors = [cmap(i / max(1, len(dsets) - 1)) if cmap is plt.cm.viridis
              else cmap(i % 10) for i in range(len(dsets))]

    fig, (ax, axr) = plt.subplots(
        2, 1, sharex=True, figsize=(9.5, 6),
        gridspec_kw={"height_ratios": [3, 1]}, layout="constrained",
    )

    for ds, c in zip(dsets, colors):
        tau_ms = ds["tau"] * 1e3
        ax.semilogx(tau_ms, ds["G"], linestyle="none", marker=".",
                    markersize=3.5, color=c, alpha=0.8)
        t_dense = np.logspace(np.log10(ds["tau"].min()),
                              np.log10(ds["tau"].max()), 400)
        g_dense = model.func(t_dense, **{n: ds["values"][n] for n in result["names"]})
        ax.semilogx(t_dense * 1e3, g_dense, color=c, linewidth=1.4,
                    label=ds["name"])
        res = ds["resid"] / ds["sigma"] if result["weighted"] else ds["resid"]
        axr.semilogx(tau_ms, res, linestyle="none", marker=".",
                     markersize=2.5, color=c, alpha=0.8)

    ax.axhline(0, color="grey", linewidth=0.6, linestyle="--")
    ax.set_ylabel("G(τ)", fontsize=12)
    ax.grid(True, which="major", linestyle="--", linewidth=0.4, alpha=0.5)
    ax.grid(True, which="minor", linestyle=":",  linewidth=0.3, alpha=0.3)

    # Linked-parameter summary box (linked values are shared, so read ds 0)
    linked_names = [n for n in result["names"] if result["linked"].get(n)]
    box = []
    ref = dsets[0]["values"]
    referr = dsets[0]["errors"]
    for n in linked_names:
        unit = next((p.unit for p in model.params if p.name == n), "")
        tag = "(fixed)" if result["fixed"].get(n) else f"± {referr[n]:.2g}"
        box.append(f"{n} = {ref[n]:.4g} {tag} {unit}".rstrip())
    gof = (f"red. χ² = {result['red_chi2']:.3g}"
           if result["weighted"] else
           f"global R² = {1 - result['ss_res'] / _grand_ss_tot(dsets):.4f}")
    box.append(gof)
    if box:
        ax.text(0.98, 0.95, "linked:\n" + "\n".join(box),
                transform=ax.transAxes, ha="right", va="top",
                fontsize=8.5, family="monospace",
                bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.85))

    ax.legend(loc="lower left", fontsize=8, framealpha=0.85, ncol=1)
    ax.set_title(
        f"Global FCS fit — {model.name}\n"
        f"{result['n_datasets']} datasets  ·  {result['n_free']} free params  ·  "
        f"{model.formula}",
        fontsize=10,
    )

    axr.axhline(0, color="grey", linewidth=0.8)
    axr.set_ylabel("resid/σ" if result["weighted"] else "resid", fontsize=10)
    axr.set_xlabel("Lag time τ (ms)", fontsize=12)
    axr.grid(True, which="major", linestyle="--", linewidth=0.4, alpha=0.5)
    axr.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f"{x:.2g}"))

    if show:
        #plot.show was static; fcs_plottools is dynamic
        #plt.show()
        fcs_plottools.show_figure(fig, np.array([ax, axr]))
    return fig, np.array([ax, axr])


def _grand_ss_tot(dsets) -> float:
    allG = np.concatenate([d["G"] for d in dsets])
    return float(np.sum((allG - allG.mean()) ** 2)) or float("nan")


def export_global_fit(result: dict, out_source: str | Path,
                      bg_factors: Optional[dict] = None) -> Tuple[Path, Path]:
    """
    Write a global-fit report (.txt) and a combined long-format curve CSV.
    ``out_source`` provides the folder (its 'fits' sibling) and filename stem.
    """
    out_source = Path(out_source)
    model = result["model"]
    out_dir = _new_fit_dir(out_source)
    report_path = out_dir / "globalfit_report.txt"
    curve_path  = out_dir / "globalfit_curves.csv"


    L: list[str] = []
    L.append("FCS global fit report")
    L.append("=" * 64)
    L.append(f"model      : {model.name}  [{model.key}]")
    L.append(f"formula    : {model.formula}")
    L.append(f"fitted     : {datetime.now().isoformat(timespec='seconds')}")
    L.append(f"datasets   : {result['n_datasets']}")
    L.append(f"free params: {result['n_free']}   observations: {result['n_obs']}   "
             f"dof: {result['dof']}")
    L.append(f"weighted   : {'yes (σ from G_std)' if result['weighted'] else 'no'}")
    L.append(f"converged  : {result['success']}  ({result['message']})")
    if result["weighted"]:
        L.append(f"red. chi^2 : {result['red_chi2']:.6g}")
    L.append("")
    L.append("Parameter linking")
    L.append("-" * 64)
    for n in result["names"]:
        state = ("fixed" if result["fixed"].get(n)
                 else "linked" if result["linked"].get(n) else "per-dataset")
        L.append(f"  {n:<8} : {state:<11} "
                 f"bounds [{_fmt(result['lowers'][n])}, {_fmt(result['uppers'][n])}]")
    L.append("")

    linked_names = [n for n in result["names"] if result["linked"].get(n)]
    if linked_names:
        L.append("Linked / shared parameters")
        L.append("-" * 64)
        ref, referr = result["datasets"][0]["values"], result["datasets"][0]["errors"]
        for n in linked_names:
            unit = next((p.unit for p in model.params if p.name == n), "")
            err = "—" if result["fixed"].get(n) else f"{referr[n]:.6g}"
            L.append(f"  {n:<8} = {ref[n]:.6g}  ± {err}  {unit}".rstrip())
        L.append("")

    L.append("Per-dataset results")
    L.append("-" * 64)
    for ds in result["datasets"]:
        L.append(f"[{ds['name']}]   points: {ds['n_points']}   R² = {ds['r2']:.5f}")
        cps = _cps_from_meta(ds.get("meta"))
        if cps is not None:
            L.append(f"    CPS      : Ch1 {_fmt_cps(cps['cps_ch1'])}   "
                     f"Ch2 {_fmt_cps(cps['cps_ch2'])}   "
                     f"{cps['cps_label']} {_fmt_cps(cps['cps_fit'])}")
        for n in result["names"]:
            if result["linked"].get(n):
                continue  # already reported above
            unit = next((p.unit for p in model.params if p.name == n), "")
            err = "—" if result["fixed"].get(n) else f"{ds['errors'][n]:.6g}"
            L.append(f"    {n:<8} = {ds['values'][n]:.6g}  ± {err}  {unit}".rstrip())
        L.append("")

    if bg_factors:
        win = bg_factors.get("_window_ns")
        L.append("Background correction")
        L.append("-" * 64)
        if win:
            L.append(f"  TCSPC tail window : {win[0]:.2f} – {win[1]:.2f} ns")
        L.append(f"  {'dataset':<24}{'f1':>9}{'f2':>9}{'factor':>10}")
        for ds in result["datasets"]:
            bg = bg_factors.get(ds["name"])
            if bg is None:
                continue
            L.append(f"  {ds['name']:<24}{bg['f1']:>9.4f}{bg['f2']:>9.4f}"
                     f"{bg['factor']:>10.4f}")
        L.append("  (corrected <N> = <N> × factor;  factor = (1−f1)(1−f2) "
                 "for cross, (1−fX)² for auto)")
        L.append("")
        
    L.extend(fcs_fisher.format_report_lines(
        fcs_fisher.fisher_from_covariance(
            result.get("cov"), result["free_labels"], result["weighted"])))
    L.append("")
    
    L.append("Information (analytical noise covariance)")
    L.append("-" * 64)
    for ds in result["datasets"]:
        L.append(f"[{ds['name']}]")
        _c = _cps_from_meta(ds.get("meta"))
        _T = (_c or {}).get("acq_time_s")
        _R = (_c or {}).get("cps_fit")
        if not _T or not _R:
            L.append(f"  not available — need CPS and acquisition time in the "
                     f"correlation header (got cps={_R!r}, T={_T!r})")
        else:
            try:
                nc = fcs_noise.noise_covariance(
                    model, ds["values"], ds["tau"],
                    T=float(_T), count_rate=float(_R))
                free_names = [p for p in result["names"]
                              if not result["fixed"].get(p, False)]
                info = fcs_noise.fisher_information(
                    model, ds["values"], ds["tau"], nc.sigma, free_names)
                L.extend(fcs_noise.information_report_lines(info))
            except Exception as e:
                L.append(f"  not available — {type(e).__name__}: {e}")
        L.append("")

    report_path.write_text("\n".join(L), encoding="utf-8")
    
    
    
    report_path.write_text("\n".join(L), encoding="utf-8")

    # Combined long-format curve CSV
    weighted = result["weighted"]
    header = ["dataset", "tau_s", "tau_ms", "G_data", "G_fit", "residual"]
    if weighted:
        header.append("sigma")
    with curve_path.open("w", encoding="utf-8-sig", newline="") as fh:
        fh.write(f"# FCS global fit curves — {model.key}\n")
        fh.write(f"# exported : {datetime.now().isoformat(timespec='seconds')}\n")
        fh.write(",".join(header) + "\n")
        for ds in result["datasets"]:
            for i in range(ds["n_points"]):
                row = [ds["name"],
                       f"{ds['tau'][i]:.10g}", f"{ds['tau'][i]*1e3:.10g}",
                       f"{ds['G'][i]:.10g}", f"{ds['Gfit'][i]:.10g}",
                       f"{ds['resid'][i]:.10g}"]
                if weighted:
                    row.append(f"{ds['sigma'][i]:.10g}")
                fh.write(",".join(row) + "\n")

    print(f"[globalfit] wrote {report_path}")
    print(f"[globalfit] wrote {curve_path}")

    # ── Wide parameter table (one row per dataset) for spreadsheets ───────────
    # Designed for downstream "parameter vs variable" plots (e.g. tau_D vs
    # concentration): each dataset is a row, each parameter a value+error pair.
    # ── Wide parameter table (one row per dataset) for spreadsheets ───────────
    # Calibrate Volume reads the .csv; the .xlsx is the convenient view.
    params_path = out_dir / "globalfit_params.csv"
    xlsx_path   = out_dir / "globalfit_params.xlsx"
    linked_names = [n for n in result["names"] if result["linked"].get(n)]
    fixed_names  = [n for n in result["names"] if result["fixed"].get(n)]

    p_header = ["dataset"]
    for n in result["names"]:
        p_header += [n, f"{n}_err"]
    has_N = "G0" in result["names"]
    if has_N:
        p_header += ["N", "N_err"]                 # <N> = 1/G0, propagated error
    has_bg = bool(bg_factors) and has_N
    if has_bg:
        p_header += ["bg_factor", "N_corr", "N_corr_err"]
    has_cps = any(_cps_from_meta(ds.get("meta")) is not None
                  for ds in result["datasets"])
    if has_cps:
        p_header += ["cps_ch1", "cps_ch2", "cps_fit", "acq_time_s"]
    p_header += ["r2", "n_points"]

    # Build native-typed rows ONCE, then render to both .csv and .xlsx.
    data_rows: list[list] = []
    for ds in result["datasets"]:
        row: list = [ds["name"]]
        for n in result["names"]:
            row.append(float(ds["values"][n]))
            row.append(float(ds["errors"][n]))
        N = N_err = float("nan")
        if has_N:
            g0  = ds["values"]["G0"]
            g0e = ds["errors"]["G0"]
            if g0 > 0:
                N     = 1.0 / g0
                N_err = g0e / (g0 * g0)             # σ_N = σ_G0 / G0²
            row.append(float(N))
            row.append(float(N_err))
        if has_bg:
            bg = bg_factors.get(ds["name"], {})
            factor = bg.get("factor", float("nan"))
            row.append(float(factor))
            row.append(float(N * factor))
            row.append(float(N_err * factor))       # factor treated as exact
        if has_cps:
            cps = _cps_from_meta(ds.get("meta")) or {}
            for key in ("cps_ch1", "cps_ch2", "cps_fit", "acq_time_s"):
                v = cps.get(key)
                row.append(float(v) if v is not None else float("nan"))
        row.append(float(ds["r2"]))
        row.append(int(ds["n_points"]))
        data_rows.append(row)

    # Comment/header block (no em-dash → no mojibake; Calibrate Volume ignores
    # these '#' lines except for the '# key : value' metadata it parses).
    comments: list[str] = []
    comments.append("FCS global fit - parameter table")
    comments.append(f"model : {model.name} [{model.key}]")
    comments.append(f"exported : {datetime.now().isoformat(timespec='seconds')}")
    comments.append(f"weighted : {'yes' if result['weighted'] else 'no'}")
    comments.append(f"linked : {', '.join(linked_names) if linked_names else '(none)'}")
    comments.append(f"fixed : {', '.join(fixed_names) if fixed_names else '(none)'}")
    if result["weighted"]:
        comments.append(f"global_red_chi2 : {result['red_chi2']:.6g}")
    comments.append("units : tau_D in seconds")
    if has_N:
        comments.append("note : N = 1/G0 (geometric factor 1); N_err = G0_err / G0^2")
    if has_bg:
        win = bg_factors.get("_window_ns")
        if win:
            comments.append(f"background_window_ns : {win[0]:.3f} - {win[1]:.3f}")
        comments.append("note : N_corr = N * bg_factor (background-corrected occupancy)")
    if has_cps:
        comments.append("note : cps_* are mean count rates (Hz) from the correlation "
                        "headers; cps_fit = that channel (auto) or the Ch1/Ch2 average (cross)")

    def _csv_cell(v) -> str:
        return f"{v:.10g}" if isinstance(v, float) else str(v)

    with params_path.open("w", encoding="utf-8", newline="") as fh:
        for line in comments:
            fh.write(f"# {line}\n")
        fh.write(",".join(p_header) + "\n")
        for row in data_rows:
            fh.write(",".join(_csv_cell(v) for v in row) + "\n")
    print(f"[globalfit] wrote {params_path}")

    _write_params_xlsx(xlsx_path, comments, p_header, data_rows)

    return report_path, curve_path, params_path


# ── Background correction ─────────────────────────────────────────────────────

def resolve_source_fcs(dataset: dict) -> Optional[Path]:
    """
    Locate the .fcs file a correlation CSV came from, using the 'source file'
    recorded in its header and the analysis/-folder convention.
    """
    meta = dataset.get("meta", {})
    src_name = meta.get("source file") or meta.get("source")
    csvp = Path(dataset["path"])
    base = csvp.parent.parent if csvp.parent.name.lower() == "analysis" else csvp.parent

    candidates: list[Path] = []
    if src_name:
        candidates += [base / src_name, csvp.parent / src_name]
    stem = csvp.stem
    if "_correlation" in stem:                      # '<stem>_correlation_<type>'
        gstem = stem.split("_correlation")[0]
        candidates += [base / f"{gstem}.fcs", csvp.parent / f"{gstem}.fcs"]
    for c in candidates:
        if c.exists():
            return c
    return None


def background_fraction(d: FCSData, channel: int,
                        lo_ns: float, hi_ns: float) -> float:
    """
    Estimate the uncorrelated-background fraction f = B/(F+B) for one channel
    from a TCSPC tail window where fluorescence has decayed.

    The background is flat across the laser period, so extrapolating the
    window's photon density to the full period gives the total background:

        f = (photons in window / total) × (period / window width)
    """
    micro_ns = d.ch1_micro_ns if channel == 1 else d.ch2_micro_ns
    total = len(micro_ns)
    if total == 0 or hi_ns <= lo_ns:
        return 0.0
    period = d.laser_period_ns
    n_win = int(np.count_nonzero((micro_ns >= lo_ns) & (micro_ns < hi_ns)))
    f = (n_win / total) * (period / (hi_ns - lo_ns))
    return float(np.clip(f, 0.0, 1.0))


def default_background_window(d: FCSData) -> Tuple[float, float]:
    """A safe tail window (ns) past the fluorescence decay for the gate default."""
    period = d.laser_period_ns
    try:
        bt, c1 = d.lifetime_histogram(channel=1, n_bins=256)
        _,  c2 = d.lifetime_histogram(channel=2, n_bins=256)
        peak_ns = float(bt[int(np.argmax(c1.astype(float) + c2.astype(float)))])
    except Exception:
        peak_ns = 0.0
    lo = max(0.60 * period, peak_ns + 15.0)
    hi = 0.97 * period
    if lo >= hi:
        lo, hi = 0.70 * period, 0.99 * period
    return float(lo), float(hi)


def compute_background_factors(datasets: list,
                               lo_ns: float, hi_ns: float) -> dict:
    """
    For each dataset, load its source .fcs and compute the amplitude
    background-correction factor over the tail window [lo_ns, hi_ns):

        cross   : factor = (1 − f1)(1 − f2)
        auto_chX: factor = (1 − fX)²

    Corrected occupancy is  <N>_corr = <N>_meas × factor.

    Returns {name: {factor, f1, f2, type, source}}.  Datasets whose source
    .fcs cannot be located get factor = nan.
    """
    cache: dict = {}
    out: dict = {}
    for ds in datasets:
        name = ds["name"]
        ctype = (ds.get("meta", {}).get("type") or "cross").strip()
        src = resolve_source_fcs(ds)
        if src is None:
            out[name] = {"factor": float("nan"), "f1": float("nan"),
                         "f2": float("nan"), "type": ctype, "source": None}
            continue
        key = str(src)
        if key not in cache:
            cache[key] = read_fcs(src)
        d = cache[key]
        f1 = background_fraction(d, 1, lo_ns, hi_ns)
        f2 = background_fraction(d, 2, lo_ns, hi_ns)
        if ctype == "auto_ch1":
            factor = (1.0 - f1) ** 2
        elif ctype == "auto_ch2":
            factor = (1.0 - f2) ** 2
        else:
            factor = (1.0 - f1) * (1.0 - f2)
        out[name] = {"factor": float(factor), "f1": f1, "f2": f2,
                     "type": ctype, "source": src}
    return out


# ── GUI workflow ──────────────────────────────────────────────────────────────

def run_model_dialog(fcs_data, parent=None, workspace_order=None):
    """
    Top-level entry: choose which data type to model, then dispatch.

    Correlation runs the fit workflow implemented in this module; Lifetime and
    PCH dispatch to run_lifetime_fit_dialog / run_pch_fit_dialog in the
    fcs_lifetime_fit and fcs_pch_fit modules respectively.

    ``workspace_order`` is an optional list of source .fcs file names in
    workspace order; when given, discovered correlation datasets are listed
    and reported in that order instead of alphabetically.
    """
    import tkinter as tk

    win = tk.Toplevel(parent)
    win.title("Model data")
    win.geometry("320x270")
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text="Model data",
             font=("Helvetica", 12, "bold"), pady=8).pack()
    tk.Label(win, text="Select the data type you want to fit.",
             font=("Helvetica", 9), fg="grey").pack()

    btns = tk.Frame(win, padx=20, pady=12)
    btns.pack(fill="both", expand=True)

    def _correlation():
        win.destroy()
        run_global_fit_dialog(fcs_data, parent=parent,
                              workspace_order=workspace_order)

    def _lifetime():
        win.destroy()
        fcs_lifetime_fit.run_lifetime_fit_dialog(fcs_data, parent=parent)

    def _pch():
        win.destroy()
        fcs_pch_fit.run_pch_fit_dialog(fcs_data, parent=parent)

    tk.Button(btns, text="Correlation", width=26, pady=6,
              command=_correlation).pack(pady=4)
    tk.Button(btns, text="Lifetime", width=26, pady=6,
              command=_lifetime).pack(pady=4)
    tk.Button(btns, text="PCH", width=26, pady=6,
              command=_pch).pack(pady=4)

    tk.Button(win, text="Cancel", width=10, command=win.destroy,
              pady=4).pack(pady=(0, 10))

    win.wait_window()


def run_global_fit_dialog(fcs_data, parent=None, workspace_order=None):
    """
    Entry point for correlation modelling: select datasets, choose a model,
    set linking / guesses / bounds, then fit, plot and export.

    Works for one dataset (a plain single-curve fit) or many (global fit with
    linked parameters).  ``workspace_order`` (source .fcs file names in
    workspace order) orders the dataset list and the output rows to match.
    """
    start_dir = fcs_data.filepath.parent
    analysis  = start_dir / "analysis"
    init_dir  = analysis if analysis.exists() else start_dir

    def _after_datasets(loaded):
        def _after_model(model):
            _global_setup_dialog(parent, model, loaded, fcs_data.filepath)
        _select_model_dialog(parent, _after_model)

    _select_datasets_dialog(parent, init_dir, _after_datasets,
                            order=workspace_order)


def _discover_correlation_csvs(folder: Path) -> list:
    """
    Return CSVs in ``folder`` that parse as correlation exports (have a lag
    axis and a G column), regardless of filename.  Robust to custom names.
    """
    folder = Path(folder)
    found = []
    if folder.exists():
        for p in sorted(folder.glob("*.csv")):
            try:
                load_correlation_csv(p)
                found.append(p)
            except Exception:
                continue
    return found


def _source_name_of_csv(path: Path) -> str:
    """Return the originating .fcs file name recorded in a correlation CSV header."""
    try:
        _, _, _, _, meta = load_correlation_csv(path)
        return (meta.get("source file") or meta.get("source") or "").strip()
    except Exception:
        return ""


def _order_paths_by_workspace(paths: list, order_names: list) -> list:
    """
    Sort correlation CSV *paths* to match the workspace file order.

    Each CSV is matched to its source .fcs (via the 'source file' header) and
    ordered by that file's position in *order_names*.  Datasets whose source
    is not in the workspace (e.g. files added by hand) are kept after the
    workspace ones, ordered alphabetically.
    """
    order_index = {name: i for i, name in enumerate(order_names)}
    src = {p: _source_name_of_csv(p) for p in paths}
    return sorted(
        paths,
        key=lambda p: (order_index.get(src[p], len(order_names)), p.name.lower()),
    )


def _select_datasets_dialog(parent, init_dir, on_done, order=None):
    """Screen — include/exclude correlation CSVs (auto-discovered + add more)."""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    init_dir = Path(init_dir)
    path_list = _discover_correlation_csvs(init_dir)
    if order:
        path_list = _order_paths_by_workspace(path_list, order)

    win = tk.Toplevel(parent)
    win.title("Select correlation datasets")
    win.geometry("540x430")
    win.grab_set()

    tk.Label(win, text="Select datasets to include in the fit",
             font=("Helvetica", 12, "bold"), pady=6).pack()
    tk.Label(win, text="Highlighted rows are included — click a row to toggle.",
             font=("Helvetica", 9), fg="grey").pack()

    lb_frame = tk.Frame(win)
    lb_frame.pack(fill="both", expand=True, padx=12, pady=6)
    scroll = tk.Scrollbar(lb_frame, orient="vertical")
    listbox = tk.Listbox(lb_frame, selectmode="multiple",
                         yscrollcommand=scroll.set, activestyle="none",
                         font=("Courier", 9))
    scroll.config(command=listbox.yview)
    scroll.pack(side="right", fill="y")
    listbox.pack(side="left", fill="both", expand=True)

    info = tk.StringVar(value="")

    def _populate(select_all=True, keep=None):
        keep = keep or set()
        listbox.delete(0, tk.END)
        for i, p in enumerate(path_list):
            listbox.insert(tk.END, p.name)
            if select_all or p in keep:
                listbox.selection_set(i)
        _update_info()

    def _update_info(*_):
        info.set(f"{len(listbox.curselection())} of {len(path_list)} included")

    listbox.bind("<<ListboxSelect>>", _update_info)
    _populate(select_all=True)

    def _add_files():
        new = filedialog.askopenfilenames(
            title="Add correlation CSV files",
            initialdir=str(init_dir),
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"),
                       ("Correlation CSV", "*correlation*.csv"),
                       ("All files", "*.*")],
            parent=win)
        if not new:
            return
        keep = {path_list[i] for i in listbox.curselection()}
        for n in new:
            pn = Path(n)
            if pn not in path_list:
                path_list.append(pn)
                keep.add(pn)
        _populate(select_all=False, keep=keep)

    if not path_list:
        info.set("No correlation CSVs found in the analysis folder — use 'Add files…'.")

    tk.Label(win, textvariable=info, font=("Helvetica", 9), fg="grey").pack()

    btns = tk.Frame(win)
    btns.pack(pady=8)
    tk.Button(btns, text="Add files…", command=_add_files,
              width=12, pady=4).pack(side="left", padx=6)

    def _next():
        sel = listbox.curselection()
        if not sel:
            messagebox.showinfo("No datasets",
                                "Select at least one dataset (click to highlight).",
                                parent=win)
            return
        chosen = [path_list[i] for i in sel]
        loaded, errors = [], []
        for p in chosen:
            try:
                tau, G, Gstd, _cols, meta = load_correlation_csv(p)
                loaded.append({"name": p.stem, "path": p, "meta": meta,
                               "tau": tau, "G": G, "sigma": Gstd})
            except Exception as e:
                errors.append(f"{p.name}: {e}")
        if errors:
            messagebox.showerror("Some files could not be read",
                                 "\n".join(errors), parent=win)
        if not loaded:
            return
        win.destroy()
        on_done(loaded)

    tk.Button(btns, text="Next →", command=_next,
              width=12, pady=4).pack(side="left", padx=6)
    tk.Button(btns, text="Cancel", command=win.destroy,
              width=10, pady=4).pack(side="left", padx=6)

    win.wait_window()


def _global_setup_dialog(parent, model, datasets, out_source):
    """Screen — per-parameter link / guess / bounds / fix, then fit."""
    import tkinter as tk
    from tkinter import messagebox

    D = len(datasets)
    all_have_sigma = all(
        ds["sigma"] is not None and np.isfinite(ds["sigma"]).any()
        for ds in datasets
    )

    win = tk.Toplevel(parent)
    win.title(f"Global fit — {model.name}")
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text=f"Global fit — {model.name}",
             font=("Helvetica", 12, "bold"), pady=6).pack()
    tk.Label(win, text=f"{D} dataset{'s' if D != 1 else ''}  ·  {model.formula}",
             font=("Helvetica", 9), fg="grey").pack(pady=(0, 6))

    guesses0 = combined_guess(model, datasets)

    table = tk.Frame(win, padx=12, pady=4)
    table.pack(fill="x")
    for c, h in enumerate(["Parameter", "Link", "Guess", "Lower", "Upper", "Fix"]):
        tk.Label(table, text=h, font=("Helvetica", 10, "bold")).grid(
            row=0, column=c, padx=4, pady=(0, 4))

    link_vars: Dict[str, tk.BooleanVar] = {}
    guess_vars: Dict[str, tk.StringVar] = {}
    lower_vars: Dict[str, tk.StringVar] = {}
    upper_vars: Dict[str, tk.StringVar] = {}
    fixed_vars: Dict[str, tk.BooleanVar] = {}

    for r, p in enumerate(model.params, start=1):
        label = p.name + (f" ({p.unit})" if p.unit else "")
        tk.Label(table, text=label, anchor="w", width=12).grid(
            row=r, column=0, sticky="w", padx=4, pady=2)

        lkv = tk.BooleanVar(value=p.link_default if D > 1 else False)
        gv  = tk.StringVar(value=f"{guesses0.get(p.name, p.default):.6g}")
        lv  = tk.StringVar(value=_fmt(p.lower))
        uv  = tk.StringVar(value=_fmt(p.upper))
        fv  = tk.BooleanVar(value=p.fixed)

        tk.Checkbutton(table, variable=lkv,
                       state="normal" if D > 1 else "disabled").grid(row=r, column=1, padx=4)
        tk.Entry(table, textvariable=gv, width=12).grid(row=r, column=2, padx=4)
        tk.Entry(table, textvariable=lv, width=10).grid(row=r, column=3, padx=4)
        tk.Entry(table, textvariable=uv, width=10).grid(row=r, column=4, padx=4)
        tk.Checkbutton(table, variable=fv).grid(row=r, column=5, padx=4)

        link_vars[p.name]  = lkv
        guess_vars[p.name] = gv
        lower_vars[p.name] = lv
        upper_vars[p.name] = uv
        fixed_vars[p.name] = fv

    note = ("Linked = one shared value across all datasets; "
            "unlinked = an independent value per dataset.")
    if D == 1:
        note = "Linking applies with 2+ datasets (only one selected)."
    tk.Label(win, text=note, font=("Helvetica", 9), fg="grey",
             wraplength=440, justify="left").pack(fill="x", padx=12, pady=(4, 0))

    weight_var = tk.BooleanVar(value=all_have_sigma)
    tk.Checkbutton(
        win,
        text="Weight by σ (G_std)" if all_have_sigma
        else "Weight by σ — unavailable (a dataset lacks G_std)",
        variable=weight_var, anchor="w",
        state="normal" if all_have_sigma else "disabled",
    ).pack(fill="x", padx=12, pady=(6, 0))

    bg_var = tk.BooleanVar(value=False)
    tk.Checkbutton(
        win,
        text="Correct for background  (pick a TCSPC tail range)",
        variable=bg_var, anchor="w",
    ).pack(fill="x", padx=12, pady=(0, 0))

    btns = tk.Frame(win)
    btns.pack(pady=10)

    def _do_fit():
        try:
            guesses = {n: float(guess_vars[n].get()) for n in guess_vars}
            lowers  = {n: _parse_bound(lower_vars[n].get(), -np.inf) for n in lower_vars}
            uppers  = {n: _parse_bound(upper_vars[n].get(),  np.inf) for n in upper_vars}
        except ValueError:
            messagebox.showerror("Invalid input",
                                 "Guesses and bounds must be numbers "
                                 "(use 'inf' / '-inf').", parent=win)
            return
        linked = {n: (link_vars[n].get() if D > 1 else False) for n in link_vars}
        fixed  = {n: fixed_vars[n].get() for n in fixed_vars}
        for n in guesses:
            if lowers[n] >= uppers[n]:
                messagebox.showerror("Invalid bounds",
                                     f"For '{n}', lower must be < upper.", parent=win)
                return
        weighted = all_have_sigma and weight_var.get()

        # ── Optional background correction ────────────────────────────────────
        bg_factors = None
        if bg_var.get():
            # Find a dataset whose source .fcs we can load to show the histogram.
            ref_fcs = None
            for ds in datasets:
                src = resolve_source_fcs(ds)
                if src is not None:
                    try:
                        ref_fcs = read_fcs(src)
                        break
                    except Exception:
                        continue
            if ref_fcs is None:
                messagebox.showerror(
                    "Background correction unavailable",
                    "Could not locate the source .fcs file for any selected "
                    "dataset (needed for the lifetime histogram).  Keep the "
                    "correlation CSVs in their 'analysis' folder next to the "
                    ".fcs files, or untick background correction.",
                    parent=win)
                return
            lo0, hi0 = default_background_window(ref_fcs)
            window = fcs_lifetime.select_gate(
                ref_fcs, initial_gate=(lo0, hi0),
                title="Background range — TCSPC tail",
                gate_label="Background",
                confirm_text="Use this background range")
            if window is None:
                return   # user cancelled — leave the setup open
            lo_ns, hi_ns = window
            try:
                bg_factors = compute_background_factors(datasets, lo_ns, hi_ns)
            except Exception as e:
                messagebox.showerror("Background correction failed", str(e), parent=win)
                return
            bg_factors["_window_ns"] = (lo_ns, hi_ns)

        try:
            result = fit_global(model, datasets, linked, guesses,
                                lowers, uppers, fixed, weighted=weighted)
        except Exception as e:
            messagebox.showerror("Fit failed", str(e), parent=win)
            return

        win.destroy()
        report_path, _curve, _params = export_global_fit(
            result, out_source, bg_factors=bg_factors)
        fig, _axes = plot_global_fit(result, show=False)
        try:
            fig.savefig(report_path.with_suffix(".png"), dpi=150)
        except Exception as e:
            print(f"[globalfit] could not save figure: {e}")

        linked_lines = [
            f"{n} = {result['datasets'][0]['values'][n]:.4g} (linked)"
            for n in result["names"] if result["linked"].get(n)
        ]
        gof = (f"red. χ² = {result['red_chi2']:.3g}"
               if result["weighted"] else f"{result['n_datasets']} datasets")
        messagebox.showinfo(
            "Global fit complete",
            f"{model.name}\n\n"
            + ("\n".join(linked_lines) if linked_lines else "(no linked parameters)")
            + f"\n\n{gof}\n\nResults saved to:\n{report_path.parent}",
            parent=parent,
        )
        #plot.show was static; fcs_plottools is dynamic
        #plt.show()
        #fcs_plottools.show_figure(fig, np.array([ax, axr]))
        fcs_plottools.show_figure(fig, _axes)
        
    tk.Button(btns, text="Fit", width=12, command=_do_fit, pady=4).pack(side="left", padx=6)
    tk.Button(btns, text="Cancel", width=10, command=win.destroy, pady=4).pack(side="left", padx=6)

    win.wait_window()


def _select_model_dialog(parent, on_choose):
    """Screen 1 — choose a model from the registry (extensible)."""
    import tkinter as tk

    win = tk.Toplevel(parent)
    win.title("Fit — select model")
    win.geometry("525x400")
    win.minsize(440,360)
    win.resizable(True, True)
    win.grab_set()

    tk.Label(win, text="Select a fit model",
             font=("Helvetica", 12, "bold"), pady=8).pack()

    models = fcs_models.list_models()
    key_var = tk.StringVar(value=models[0].key)

    list_frame = tk.LabelFrame(win, text="Models", padx=10, pady=6)
    list_frame.pack(fill="x", padx=12, pady=(0, 6))
    for m in models:
        tk.Radiobutton(list_frame, text=m.name, variable=key_var,
                       value=m.key, anchor="w").pack(fill="x")

    desc = tk.Text(win, height=8, wrap="word", font=("Courier", 9),
                   bg="#f7f7f7", relief="flat", padx=8, pady=6)
    desc.pack(fill="both", expand=True, padx=12, pady=(0, 6))

    def _refresh_desc(*_):
        m = fcs_models.get_model(key_var.get())
        desc.config(state="normal")
        desc.delete("1.0", tk.END)
        desc.insert(tk.END, m.description)
        desc.config(state="disabled")

    key_var.trace_add("write", _refresh_desc)
    _refresh_desc()

    btns = tk.Frame(win)
    btns.pack(pady=8)

    def _next():
        m = fcs_models.get_model(key_var.get())
        win.destroy()
        on_choose(m)

    tk.Button(btns, text="Next →", width=12, command=_next, pady=4).pack(side="left", padx=6)
    tk.Button(btns, text="Cancel", width=10, command=win.destroy, pady=4).pack(side="left", padx=6)

    win.wait_window()


def _parse_bound(text: str, default: float) -> float:
    """Parse a bound entry, accepting blank, 'inf', '-inf'."""
    t = text.strip().lower()
    if t in ("", "inf", "+inf", "infinity"):
        return np.inf if t != "" else default
    if t in ("-inf", "-infinity"):
        return -np.inf
    return float(t)


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python fcs_fit.py <correlation.csv> [model_key]")
        print("Available models:")
        for m in fcs_models.list_models():
            print(f"  {m.key:<22} {m.name}")
        sys.exit(1)

    path = Path(sys.argv[1])
    key  = sys.argv[2] if len(sys.argv) > 2 else fcs_models.list_models()[0].key
    model = fcs_models.get_model(key)

    tau_s, G, G_std, _cols, _meta = load_correlation_csv(path)
    guesses = auto_guess(model, tau_s, G)
    lowers = {p.name: p.lower for p in model.params}
    uppers = {p.name: p.upper for p in model.params}
    fixed  = {p.name: p.fixed for p in model.params}

    result = fit_global(model, datasets, linked, guesses,
                                lowers, uppers, fixed, weighted=weighted)
    export_global_fit(result, path)
    plot_global_fit(result, path.name)
